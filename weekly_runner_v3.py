"""
Zentra Cloud – Weekly Automated Runner  (with Open-Meteo Forecast)
==================================================================
Runs on a scheduled cadence via Windows Task Scheduler.
Fetches the most recent 7 complete days of weather sensor data,
exports it to Excel, runs a quality check, and emails
a formatted report that includes:
  • A daily summary of recent Zentra sensor readings
  • A 7-day ahead weather forecast from Open-Meteo

Requirements:
    pip install requests pandas openpyxl matplotlib

GitHub Actions setup:
    Store private values as repository secrets instead of committing them:
      ZENTRA_API_TOKEN
      GMAIL_APP_PASSWORD  (only needed when sending email)
"""

import json
import os
import re
import sys
import time
import smtplib
import traceback
import urllib.request
import requests
import pandas as pd
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
from io                     import StringIO
from datetime               import datetime, timedelta
from pathlib                import Path
from email.mime.image       import MIMEImage
from email.mime.multipart   import MIMEMultipart
from email.mime.text        import MIMEText
from openpyxl               import Workbook
from openpyxl.styles        import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils         import get_column_letter

try:
    import ee
except ImportError:
    ee = None


# ══════════════════════════════════════════════════════════════════════════════
#  ✏️  EDIT YOUR INPUTS HERE
# ══════════════════════════════════════════════════════════════════════════════

# ── Email ─────────────────────────────────────────────────────────────────────
GMAIL_SENDER       = os.getenv("GMAIL_SENDER", "dialameh.babak@gmail.com")
GMAIL_APP_PASSWORD = os.getenv("GMAIL_APP_PASSWORD", "")
EMAIL_RECIPIENT    = [
    #"dialameh.1@osu.edu",
    #"dialameh.babak@gmail.com",
    # "colleague@example.com",   # ← add more recipients here
]

# ── Sensor / API ──────────────────────────────────────────────────────────────
API_TOKEN  = os.getenv("ZENTRA_API_TOKEN", "Token 2cc62226cf6674b8501cd799f90dc32e1c6052f5")
DEVICE_SN  = os.getenv("ZENTRA_DEVICE_SN", "z6-27971")
PORT_NUM   = int(os.getenv("ZENTRA_PORT_NUM", "6"))

BASE_URL        = "https://zentracloud.com/api/v4/get_readings/"
PER_PAGE        = 500
RATE_LIMIT_WAIT = 62
OUT_DIR         = os.getenv(
    "ZENTRA_OUT_DIR",
    r"C:\Users\dialameh.1\OneDrive - The Ohio State University\ZentraCloud",
)

# ── Location (used for Open-Meteo forecast) ───────────────────────────────────
LATITUDE  = float(os.getenv("SENSOR_LATITUDE", "40.743377"))   # Decimal degrees, -90  to  90
LONGITUDE = float(os.getenv("SENSOR_LONGITUDE", "-84.024581"))  # Decimal degrees, -180 to 180

# ── Google Earth Engine / Sentinel-2 NDVI ─────────────────────────────────────
GEE_FIELD_ASSET_ID = os.getenv("GEE_FIELD_ASSET_ID", "users/dialamehbabak/Test")
GEE_PROJECT = os.getenv("GEE_PROJECT") or os.getenv("GOOGLE_CLOUD_PROJECT")
DEFAULT_GEE_SERVICE_ACCOUNT = "gee-dashboard-runner@babakdialameh-92.iam.gserviceaccount.com"
GEE_SERVICE_ACCOUNT = os.getenv("GEE_SERVICE_ACCOUNT", "")
GEE_SERVICE_ACCOUNT_KEY_JSON = os.getenv("GEE_SERVICE_ACCOUNT_KEY_JSON", "")
GEE_SERVICE_ACCOUNT_KEY_FILE = os.getenv("GEE_SERVICE_ACCOUNT_KEY_FILE", "")
NDVI_IMAGE_FILENAME = "sentinel2_ndvi_latest.png"

# ══════════════════════════════════════════════════════════════════════════════

EMAIL_SUMMARY_EXCLUDE = {"xaxislevel", "yaxislevel"}


# ── Date range: rolling 7-day history ending yesterday ───────────────────────
def historical_range():
    today = datetime.now().date()
    end_date = today - timedelta(days=1)
    start_date = end_date - timedelta(days=6)
    return start_date.strftime("%Y-%m-%d"), end_date.strftime("%Y-%m-%d")


def zentra_fetch_end_date():
    # Zentra's end_date behaves like an exclusive UTC-oriented boundary here.
    # Fetch through today to include the full local Eastern day for yesterday,
    # then filter processed rows back to the displayed range.
    return datetime.now().date().strftime("%Y-%m-%d")


# ══════════════════════════════════════════════════════════════════════════════
#  ZENTRA CLOUD  —  API helpers
# ══════════════════════════════════════════════════════════════════════════════

def build_session():
    token = API_TOKEN.strip()
    if not token:
        raise ValueError(
            "Missing Zentra API token. Set the ZENTRA_API_TOKEN environment variable "
            "or add it as a GitHub repository secret."
        )
    if not token.lower().startswith("token "):
        token = f"Token {token}"
    s = requests.Session()
    s.headers.update({"Authorization": token})
    return s


def parse_data(raw_data) -> pd.DataFrame:
    if isinstance(raw_data, str):
        try:
            parsed = json.loads(raw_data)
        except json.JSONDecodeError:
            return pd.DataFrame()
        if isinstance(parsed, dict) and "columns" in parsed:
            return pd.DataFrame(parsed.get("data", []), columns=parsed["columns"])
        if isinstance(parsed, list):
            return pd.DataFrame(parsed)
        return pd.DataFrame()
    if isinstance(raw_data, dict):
        frames = []
        for sensor_name, sensor_block in raw_data.items():
            if isinstance(sensor_block, dict) and "readings" in sensor_block:
                df = pd.DataFrame(sensor_block["readings"])
                df.insert(0, "sensor_name", sensor_name)
                frames.append(df)
            elif isinstance(sensor_block, list):
                df = pd.DataFrame(sensor_block)
                df.insert(0, "sensor_name", sensor_name)
                frames.append(df)
        return pd.concat(frames, ignore_index=True) if frames else pd.DataFrame()
    if isinstance(raw_data, list):
        return pd.DataFrame(raw_data)
    return pd.DataFrame()


def date_chunks(start: str, end: str, days: int = 7):
    """Yield Zentra API chunks using an exclusive end-date boundary."""
    cur = datetime.strptime(start, "%Y-%m-%d")
    fin = datetime.strptime(end,   "%Y-%m-%d")
    while cur < fin:
        nxt = min(cur + timedelta(days=days), fin)
        yield cur.strftime("%Y-%m-%d"), nxt.strftime("%Y-%m-%d")
        cur = nxt


def fetch_chunk(session, chunk_start, chunk_end, chunk_num) -> pd.DataFrame:
    chunk_frames = []
    page = 1
    while True:
        params = {
            "device_sn":     DEVICE_SN,
            "start_date":    chunk_start,
            "end_date":      chunk_end,
            "output_format": "df",
            "sort_by":       "asc",
            "per_page":      PER_PAGE,
            "page_num":      page,
        }
        r = session.get(BASE_URL, params=params, timeout=60)
        if r.status_code == 429:
            print(f"    ⚠  Rate-limited. Waiting {RATE_LIMIT_WAIT}s …")
            time.sleep(RATE_LIMIT_WAIT)
            r = session.get(BASE_URL, params=params, timeout=60)
        if not r.ok:
            print(f"    ✗ HTTP {r.status_code}: {r.reason}")
            break

        envelope    = r.json()
        pagination  = envelope.get("pagination", {})
        total_pages = pagination.get("total_pages", 1)
        print(f"    page {page}/{total_pages}", end=" ", flush=True)

        df_page = parse_data(envelope.get("data"))
        if not df_page.empty:
            chunk_frames.append(df_page)

        if page >= total_pages:
            break
        page += 1
        print(f"⏳{RATE_LIMIT_WAIT}s…", end=" ", flush=True)
        time.sleep(RATE_LIMIT_WAIT)

    print()
    return pd.concat(chunk_frames, ignore_index=True) if chunk_frames else pd.DataFrame()


def fetch_all(start_date: str, end_date: str) -> pd.DataFrame:
    session    = build_session()
    all_frames = []
    # Zentra can silently omit the newest local day from a 7-day API chunk.
    # Six-day chunks keep the rolling dashboard range complete after filtering.
    chunks     = list(date_chunks(start_date, end_date, days=6))

    for i, (cs, ce) in enumerate(chunks, start=1):
        print(f"  Chunk {i}/{len(chunks)}: {cs} → {ce}")
        df_chunk = fetch_chunk(session, cs, ce, chunk_num=i)
        if not df_chunk.empty:
            all_frames.append(df_chunk)
        if i < len(chunks):
            print("  Continuing to next date chunk …")

    if not all_frames:
        return pd.DataFrame()

    combined = pd.concat(all_frames, ignore_index=True)
    combined.drop_duplicates(inplace=True)

    if "timestamp_utc" in combined.columns:
        combined["datetime_utc"] = pd.to_datetime(
            combined["timestamp_utc"], unit="s", utc=True, errors="coerce"
        )
    elif "datetime" in combined.columns:
        combined["datetime_utc"] = pd.to_datetime(
            combined["datetime"], utc=True, errors="coerce"
        )

    return combined


# ══════════════════════════════════════════════════════════════════════════════
#  ZENTRA CLOUD  —  Processing
# ══════════════════════════════════════════════════════════════════════════════

def to_title(api_name: str) -> str:
    return api_name.replace("_", " ").title()


def normalize_col_name(name: str) -> str:
    return re.sub(r"[^a-z0-9]+", "", str(name).lower())


def process(df: pd.DataFrame) -> tuple[pd.DataFrame, dict, bool, int, str | None]:
    """Filter port, average hourly, pivot. Returns (df_out, quality_report,
    delayed_start, delayed_hours, delayed_start_str)."""

    # Port filter
    if PORT_NUM is not None and "port_num" in df.columns:
        df = df[df["port_num"] == PORT_NUM].reset_index(drop=True)

    # Keep needed columns
    keep = [c for c in ["datetime", "measurement", "value"] if c in df.columns]
    df   = df[keep].copy()
    df["value"] = pd.to_numeric(df["value"], errors="coerce")

    # Hourly average (Eastern Time)
    df["datetime"] = pd.to_datetime(df["datetime"], utc=True, errors="coerce")
    df["datetime"] = df["datetime"].dt.tz_convert("America/New_York")
    df["datetime"] = df["datetime"].dt.floor("h")
    df = (
        df.groupby(["datetime", "measurement"], sort=True)["value"]
        .mean().round(4).reset_index()
    )

    # Pivot
    detected   = sorted(df["measurement"].dropna().unique().tolist())
    df_pivot   = df.pivot(index="datetime", columns="measurement", values="value")
    df_pivot.columns.name = None
    df_pivot.reset_index(inplace=True)
    rename_map = {m: to_title(m) for m in detected}
    df_pivot.rename(columns=rename_map, inplace=True)

    # Quality check
    data_cols      = [c for c in df_pivot.columns if c != "datetime"]
    quality_report = {}
    for col in data_cols:
        missing = int(df_pivot[col].isna().sum())
        quality_report[col] = missing

    # Detect delayed start
    actual_start        = df_pivot["datetime"].min()
    actual_start_naive  = pd.Timestamp(actual_start).tz_localize(None)
    expected_mon_00     = actual_start.floor("D") - pd.Timedelta(days=actual_start.dayofweek)
    expected_mon_00     = expected_mon_00.tz_localize(None) if expected_mon_00.tzinfo else expected_mon_00
    delayed_start       = actual_start_naive > expected_mon_00
    delayed_hours       = int((actual_start_naive - expected_mon_00).total_seconds() / 3600) if delayed_start else 0
    delayed_start_str   = actual_start_naive.strftime("%Y-%m-%d %H:%M") if delayed_start else None

    # Format datetime
    df_pivot["datetime"] = df_pivot["datetime"].dt.tz_localize(None).dt.strftime("%Y-%m-%d %H:%M")
    df_pivot.rename(columns={"datetime": "Datetime"}, inplace=True)

    return df_pivot, quality_report, delayed_start, delayed_hours, delayed_start_str


def build_daily_summary(df_out: pd.DataFrame) -> pd.DataFrame:
    """
    Aggregate the hourly Zentra pivot table into a per-day summary.
    Returns a wide DataFrame with Date as the first column and each sensor
    variable as a column.
    """
    df = df_out.copy()
    df["Date"] = pd.to_datetime(df["Datetime"]).dt.date
    sensor_cols = [
        c for c in df.columns
        if c not in ("Datetime", "Date")
        and normalize_col_name(c) not in EMAIL_SUMMARY_EXCLUDE
    ]

    if not sensor_cols:
        return pd.DataFrame(columns=["Date"])

    summary = df.groupby("Date", as_index=False)[sensor_cols].mean(numeric_only=True)
    summary["Date"] = summary["Date"].astype(str)
    summary[sensor_cols] = summary[sensor_cols].round(2)
    return summary[["Date", *sensor_cols]]


def filter_output_to_date_range(df_out: pd.DataFrame, start_date: str, end_date: str) -> pd.DataFrame:
    """Keep processed hourly rows within the displayed local-date range."""
    df = df_out.copy()
    dates = pd.to_datetime(df["Datetime"], errors="coerce").dt.date
    start = datetime.strptime(start_date, "%Y-%m-%d").date()
    end = datetime.strptime(end_date, "%Y-%m-%d").date()
    return df[(dates >= start) & (dates <= end)].reset_index(drop=True)


# ══════════════════════════════════════════════════════════════════════════════
#  OPEN-METEO  —  7-Day Forecast
# ══════════════════════════════════════════════════════════════════════════════

WMO_CODES = {
    0:  "Clear sky",       1:  "Mainly clear",      2:  "Partly cloudy",
    3:  "Overcast",        45: "Foggy",              48: "Icy fog",
    51: "Light drizzle",   53: "Moderate drizzle",  55: "Dense drizzle",
    61: "Slight rain",     63: "Moderate rain",     65: "Heavy rain",
    71: "Slight snow",     73: "Moderate snow",     75: "Heavy snow",
    77: "Snow grains",     80: "Slight showers",    81: "Moderate showers",
    82: "Violent showers", 85: "Slight snow showers", 86: "Heavy snow showers",
    95: "Thunderstorm",    96: "Thunderstorm w/ hail", 99: "Thunderstorm w/ heavy hail",
}


def describe_weather(code) -> str:
    return WMO_CODES.get(int(code), f"Code {int(code)}")


def fetch_forecast(start_date: str, days: int = 7) -> list[dict] | None:
    """
    Fetch a forecast from Open-Meteo starting at start_date.
    Returns a list of daily dicts, or None on failure.
    """
    from datetime import datetime as dt, timedelta as td
    end_date = (dt.strptime(start_date, "%Y-%m-%d") + td(days=days - 1)).strftime("%Y-%m-%d")

    daily_vars = [
        "temperature_2m_max",
        "temperature_2m_min",
        "precipitation_sum",
        "precipitation_probability_max",
        "windspeed_10m_max",
        "weathercode",
    ]
    params = (
        f"latitude={LATITUDE}"
        f"&longitude={LONGITUDE}"
        f"&daily={','.join(daily_vars)}"
        f"&start_date={start_date}"
        f"&end_date={end_date}"
        f"&timezone=auto"
    )
    url = f"https://api.open-meteo.com/v1/forecast?{params}"
    print(f"\n  Fetching Open-Meteo forecast for {start_date} → {end_date} …")

    try:
        with urllib.request.urlopen(url, timeout=10) as response:
            data = json.loads(response.read().decode())
    except Exception as e:
        print(f"  ✗ Open-Meteo fetch failed: {e}")
        return None

    daily = data.get("daily", {})
    dates         = daily.get("time", [])
    temp_max      = daily.get("temperature_2m_max", [])
    temp_min      = daily.get("temperature_2m_min", [])
    precip        = daily.get("precipitation_sum", [])
    precip_prob   = daily.get("precipitation_probability_max", [])
    windspeed     = daily.get("windspeed_10m_max", [])
    weathercodes  = daily.get("weathercode", [])

    rows = []
    for i, date in enumerate(dates):
        rows.append({
            "Date":        date,
            "Condition":   describe_weather(weathercodes[i]) if i < len(weathercodes) else "N/A",
            "Max °C":      f"{temp_max[i]:.1f}"   if i < len(temp_max)    else "N/A",
            "Min °C":      f"{temp_min[i]:.1f}"   if i < len(temp_min)    else "N/A",
            "Rain mm":     f"{precip[i]:.1f}"     if i < len(precip)      else "N/A",
            "Rain %":      f"{int(precip_prob[i])}%" if i < len(precip_prob) else "N/A",
            "Wind km/h":   f"{windspeed[i]:.1f}"  if i < len(windspeed)   else "N/A",
        })
    print(f"  ✔  Forecast received ({len(rows)} days).")
    return rows


def find_sensor_col(df: pd.DataFrame, *names: str) -> str | None:
    lookup = {normalize_col_name(col): col for col in df.columns}
    for name in names:
        col = lookup.get(normalize_col_name(name))
        if col:
            return col
    return None


def parse_number(value) -> float | None:
    if pd.isna(value):
        return None
    text = str(value).replace("%", "").strip()
    try:
        return float(text)
    except ValueError:
        return None


def save_chart(fig, path: str, cid: str, charts: dict[str, str]) -> None:
    try:
        fig.savefig(path, bbox_inches="tight")
        charts[cid] = path
    except Exception as e:
        print(f"  ⚠  Could not save chart {os.path.basename(path)}: {e}")


def build_email_charts(
    df_out: pd.DataFrame,
    forecast_rows: list[dict] | None,
    second_week_forecast_rows: list[dict] | None,
    out_dir: str,
    timestamp: str,
) -> dict[str, str]:
    """Create inline chart PNGs for the email and return cid-to-path mapping."""
    charts = {}
    os.makedirs(out_dir, exist_ok=True)

    df = df_out.copy()
    df["Date"] = pd.to_datetime(df["Datetime"]).dt.date
    forecast = pd.DataFrame(forecast_rows or [])
    second_week_forecast = pd.DataFrame(second_week_forecast_rows or [])
    forecast = pd.concat(
        [frame for frame in [forecast, second_week_forecast] if not frame.empty],
        ignore_index=True,
    ) if (not forecast.empty or not second_week_forecast.empty) else pd.DataFrame()
    if not forecast.empty:
        forecast["Date"] = pd.to_datetime(forecast["Date"]).dt.date

    precip_col = find_sensor_col(df, "Precipitation")
    precip_frames = []
    if precip_col:
        past_precip = (
            df.groupby("Date")[precip_col]
            .sum(min_count=1)
            .reset_index(name="Precipitation (mm)")
        )
        past_precip["Period"] = "Historical"
        precip_frames.append(past_precip)
    if not forecast.empty and "Rain mm" in forecast:
        future_precip = forecast[["Date", "Rain mm"]].copy()
        future_precip["Precipitation (mm)"] = future_precip["Rain mm"].map(parse_number)
        future_precip["Period"] = "Forecast"
        precip_frames.append(future_precip[["Date", "Precipitation (mm)", "Period"]])

    if precip_frames:
        precip_data = pd.concat(precip_frames, ignore_index=True).dropna(subset=["Precipitation (mm)"])
        if not precip_data.empty:
            fig, ax = plt.subplots(figsize=(8, 3.8), dpi=140)
            colors = precip_data["Period"].map({"Historical": "#5B9BD5", "Forecast": "#70AD47"}).fillna("#A5A5A5")
            ax.bar(precip_data["Date"].astype(str), precip_data["Precipitation (mm)"], color=colors)
            ax.set_title("Precipitation: 7-Day Historical and 14-Day Forecast")
            ax.set_ylabel("mm", fontsize=13)
            ax.tick_params(axis="both", labelsize=12)
            ax.tick_params(axis="x", rotation=35)
            ax.grid(axis="y", alpha=0.25)
            ax.spines[["top", "right"]].set_visible(False)
            handles = [
                plt.Rectangle((0, 0), 1, 1, color="#5B9BD5", label="Historical"),
                plt.Rectangle((0, 0), 1, 1, color="#70AD47", label="Forecast"),
            ]
            ax.legend(handles=handles, frameon=False)
            fig.tight_layout()
            path = os.path.join(out_dir, f"zentra_precipitation_chart_{timestamp}.png")
            save_chart(fig, path, "precip_chart", charts)
            plt.close(fig)

    temp_col = find_sensor_col(df, "Air Temperature", "Temperature")
    temp_series = []
    if temp_col:
        past_temp = df.groupby("Date")[temp_col].agg(["min", "max"]).reset_index()
        temp_series.append(("Historical min", past_temp["Date"], past_temp["min"], "#5B9BD5", "--"))
        temp_series.append(("Historical max", past_temp["Date"], past_temp["max"], "#C00000", "--"))
    if not forecast.empty and {"Min °C", "Max °C"}.issubset(forecast.columns):
        future_min = forecast["Min °C"].map(parse_number)
        future_max = forecast["Max °C"].map(parse_number)
        temp_series.append(("Forecast min", forecast["Date"], future_min, "#70AD47", "-"))
        temp_series.append(("Forecast max", forecast["Date"], future_max, "#ED7D31", "-"))

    if temp_series:
        fig, ax = plt.subplots(figsize=(8, 3.8), dpi=140)
        plotted = False
        for label, dates, values, color, linestyle in temp_series:
            series = pd.DataFrame({"Date": dates, "Temperature": values}).dropna()
            if series.empty:
                continue
            ax.plot(
                series["Date"].astype(str),
                series["Temperature"],
                marker="o",
                linewidth=2,
                linestyle=linestyle,
                color=color,
                label=label,
            )
            plotted = True
        if plotted:
            ax.set_title("Min and Max Temperature: 7-Day Historical and 14-Day Forecast")
            ax.set_ylabel("°C", fontsize=13)
            ax.tick_params(axis="both", labelsize=12)
            ax.tick_params(axis="x", rotation=35)
            ax.grid(axis="y", alpha=0.25)
            ax.spines[["top", "right"]].set_visible(False)
            ax.legend(frameon=False, ncol=2)
            fig.tight_layout()
            path = os.path.join(out_dir, f"zentra_temperature_chart_{timestamp}.png")
            save_chart(fig, path, "temp_chart", charts)
        plt.close(fig)

    return charts


def js_const_array(name: str, rows: list[dict]) -> str:
    payload = json.dumps(rows, indent=2)
    return f"const {name} = {payload};"


def replace_js_const_array(js_text: str, name: str, rows: list[dict]) -> str:
    pattern = rf"const {re.escape(name)} = \[.*?\];"
    replacement = js_const_array(name, rows)
    updated, count = re.subn(pattern, replacement, js_text, count=1, flags=re.S)
    if count != 1:
        raise ValueError(f"Could not find const {name} in app.js")
    return updated


def replace_js_const_string(js_text: str, name: str, value: str) -> str:
    pattern = rf'const {re.escape(name)} = ".*?";'
    replacement = f'const {name} = {json.dumps(value)};'
    updated, count = re.subn(pattern, replacement, js_text, count=1, flags=re.S)
    if count != 1:
        raise ValueError(f"Could not find const {name} in app.js")
    return updated


def replace_js_const_value(js_text: str, name: str, value) -> str:
    pattern = rf"const {re.escape(name)} = .*?;"
    replacement = f"const {name} = {json.dumps(value)};"
    updated, count = re.subn(pattern, replacement, js_text, count=1, flags=re.S)
    if count != 1:
        raise ValueError(f"Could not find const {name} in app.js")
    return updated


def round_or_none(value, digits: int = 2):
    if value is None or pd.isna(value):
        return None
    return round(float(value), digits)


def forecast_table_rows(forecast: pd.DataFrame) -> list[dict]:
    rows = []
    if forecast.empty:
        return rows

    for _, row in forecast.iterrows():
        rows.append({
            "date": str(row.get("Date", "")),
            "condition": str(row.get("Condition", "")),
            "maxTemp": round_or_none(parse_number(row.get("Max °C")), 1),
            "minTemp": round_or_none(parse_number(row.get("Min °C")), 1),
            "rainMm": round_or_none(parse_number(row.get("Rain mm")), 1),
            "rainProbability": round_or_none(parse_number(row.get("Rain %")), 0),
            "windKmh": round_or_none(parse_number(row.get("Wind km/h")), 1),
        })
    return rows


def initialize_earth_engine() -> tuple[bool, str]:
    if ee is None:
        message = "earthengine-api is not installed; skipping Sentinel-2 NDVI."
        print(f"  ⚠  {message}")
        return False, message

    try:
        if GEE_SERVICE_ACCOUNT_KEY_JSON:
            key_json = GEE_SERVICE_ACCOUNT_KEY_JSON.replace("\\n", "\n")
            key_data = json.loads(key_json)
            service_account = (
                GEE_SERVICE_ACCOUNT
                or key_data.get("client_email")
                or DEFAULT_GEE_SERVICE_ACCOUNT
            )
            project = GEE_PROJECT or key_data.get("project_id")
            with tempfile.NamedTemporaryFile("w", suffix=".json", delete=False, encoding="utf-8") as key_file:
                key_file.write(key_json)
                key_file_path = key_file.name
            try:
                credentials = ee.ServiceAccountCredentials(service_account, key_file=key_file_path)
                ee.Initialize(credentials, project=project)
            finally:
                try:
                    os.unlink(key_file_path)
                except OSError:
                    pass
            print(f"  ✔  Google Earth Engine initialized with service account {service_account}")
            return True, ""
        if GEE_SERVICE_ACCOUNT_KEY_FILE:
            service_account = GEE_SERVICE_ACCOUNT or DEFAULT_GEE_SERVICE_ACCOUNT
            credentials = ee.ServiceAccountCredentials(service_account, key_file=GEE_SERVICE_ACCOUNT_KEY_FILE)
            ee.Initialize(credentials, project=GEE_PROJECT)
            print(f"  ✔  Google Earth Engine initialized with service account {service_account}")
            return True, ""
        else:
            ee.Initialize(project=GEE_PROJECT)
            print("  ✔  Google Earth Engine initialized")
            return True, ""
    except Exception as e:
        message = f"Could not initialize Google Earth Engine: {e}"
        print(f"  ⚠  {message}")
        return False, message


def fetch_latest_sentinel2_ndvi() -> dict:
    """Download the latest Sentinel-2 NDVI PNG for the configured field asset."""
    if not GEE_FIELD_ASSET_ID:
        message = "GEE_FIELD_ASSET_ID is not set; skipping Sentinel-2 NDVI."
        print(f"  ⚠  {message}")
        return {"statusMessage": message}
    initialized, init_message = initialize_earth_engine()
    if not initialized:
        return {"statusMessage": init_message}

    project_dir = Path(__file__).resolve().parent
    out_path = project_dir / NDVI_IMAGE_FILENAME
    end_date = datetime.utcnow().date() + timedelta(days=1)
    start_date = end_date - timedelta(days=60)

    try:
        field = ee.FeatureCollection(GEE_FIELD_ASSET_ID).geometry()
        collection = (
            ee.ImageCollection("COPERNICUS/S2_SR_HARMONIZED")
            .filterBounds(field)
            .filterDate(start_date.strftime("%Y-%m-%d"), end_date.strftime("%Y-%m-%d"))
            .filter(ee.Filter.lt("CLOUDY_PIXEL_PERCENTAGE", 80))
            .sort("system:time_start", False)
        )
        image_count = collection.size().getInfo()
        if not image_count:
            message = "No Sentinel-2 scenes found for this field in the last 60 days."
            print(f"  ⚠  {message}")
            return {"statusMessage": message}
        image = ee.Image(collection.first())
        acquired_at = ee.Date(image.get("system:time_start")).format("YYYY-MM-dd").getInfo()
        cloud_percent = image.get("CLOUDY_PIXEL_PERCENTAGE").getInfo()
        scl = image.select("SCL")
        clear_mask = (
            scl.neq(3)
            .And(scl.neq(8))
            .And(scl.neq(9))
            .And(scl.neq(10))
            .And(scl.neq(11))
        )
        ndvi = (
            image.updateMask(clear_mask)
            .normalizedDifference(["B8", "B4"])
            .rename("NDVI")
            .clip(field)
        )
        rendered = ndvi.visualize(
            min=-0.2,
            max=0.9,
            palette=["8c510a", "d8b365", "f6e8c3", "c7eae5", "5ab4ac", "01665e"],
        )
        url = rendered.getThumbURL({
            "region": field.bounds(1).coordinates().getInfo(),
            "dimensions": 900,
            "format": "png",
        })
        urllib.request.urlretrieve(url, out_path)
        print(f"  ✔  Sentinel-2 NDVI saved → {out_path}")
        return {
            "imageUrl": f"{NDVI_IMAGE_FILENAME}?v={datetime.now().strftime('%Y%m%d%H%M%S')}",
            "acquiredAt": acquired_at,
            "cloudPercent": round(float(cloud_percent), 1) if cloud_percent is not None else "",
        }
    except Exception as e:
        message = f"Sentinel-2 NDVI fetch failed: {e}"
        print(f"  ⚠  {message}")
        return {"statusMessage": message}


def update_dashboard_ndvi_metadata(ndvi_meta: dict) -> None:
    project_dir = Path(__file__).resolve().parent
    app_js_path = project_dir / "app.js"
    if not app_js_path.exists():
        print(f"  ⚠  Dashboard app.js not found: {app_js_path}")
        return

    js_text = app_js_path.read_text(encoding="utf-8")
    js_text = replace_js_const_string(js_text, "ndviImageUrl", ndvi_meta.get("imageUrl", ""))
    js_text = replace_js_const_string(js_text, "ndviAcquiredAt", ndvi_meta.get("acquiredAt", ""))
    js_text = replace_js_const_value(js_text, "ndviCloudPercent", ndvi_meta.get("cloudPercent", ""))
    js_text = replace_js_const_string(js_text, "ndviStatusMessage", ndvi_meta.get("statusMessage", ""))
    app_js_path.write_text(js_text, encoding="utf-8")


def update_dashboard_app_js(
    df_out: pd.DataFrame,
    forecast_rows: list[dict] | None,
    second_week_forecast_rows: list[dict] | None = None,
) -> None:
    """Refresh static dashboard chart/insight data in app.js."""
    project_dir = Path(__file__).resolve().parent
    app_js_path = project_dir / "app.js"
    if not app_js_path.exists():
        app_js_path = project_dir / "docs" / "app.js"
    if not app_js_path.exists():
        print(f"  ⚠  Dashboard app.js not found: {app_js_path}")
        return

    df = df_out.copy()
    df["Date"] = pd.to_datetime(df["Datetime"]).dt.date.astype(str)
    forecast = pd.DataFrame(forecast_rows or [])
    second_week_forecast = pd.DataFrame(second_week_forecast_rows or [])
    chart_forecast = pd.concat(
        [frame for frame in [forecast, second_week_forecast] if not frame.empty],
        ignore_index=True,
    ) if (not forecast.empty or not second_week_forecast.empty) else pd.DataFrame()
    temp_col = find_sensor_col(df, "Air Temperature", "Temperature")
    precip_col = find_sensor_col(df, "Precipitation")
    solar_col = find_sensor_col(df, "Solar Radiation")
    wind_col = find_sensor_col(df, "Wind Speed")

    precip_data = []
    if precip_col:
        past_precip = df.groupby("Date")[precip_col].sum(min_count=1).reset_index()
        for _, row in past_precip.iterrows():
            if pd.notna(row[precip_col]):
                precip_data.append({
                    "date": row["Date"],
                    "value": round(float(row[precip_col]), 2),
                    "type": "past",
                })

    if not chart_forecast.empty and "Date" in chart_forecast.columns and "Rain mm" in chart_forecast.columns:
        for _, row in chart_forecast.iterrows():
            rain = parse_number(row.get("Rain mm"))
            if rain is not None:
                precip_data.append({
                    "date": str(row["Date"]),
                    "value": round(float(rain), 2),
                    "type": "future",
                })

    temp_data = []
    if temp_col:
        past_temp = df.groupby("Date")[temp_col].agg(["min", "max"]).reset_index()
        for _, row in past_temp.iterrows():
            if pd.notna(row["min"]) and pd.notna(row["max"]):
                temp_data.append({
                    "date": row["Date"],
                    "pastMin": round(float(row["min"]), 2),
                    "pastMax": round(float(row["max"]), 2),
                })

    if not chart_forecast.empty and {"Date", "Min °C", "Max °C"}.issubset(chart_forecast.columns):
        for _, row in chart_forecast.iterrows():
            min_temp = parse_number(row.get("Min °C"))
            max_temp = parse_number(row.get("Max °C"))
            if min_temp is not None and max_temp is not None:
                temp_data.append({
                    "date": str(row["Date"]),
                    "futureMin": round(float(min_temp), 2),
                    "futureMax": round(float(max_temp), 2),
                })

    forecast_risk_data = []
    if not forecast.empty:
        for _, row in forecast.iterrows():
            risk_row = {
                "date": str(row.get("Date", "")),
                "rain": round(float(parse_number(row.get("Rain mm")) or 0), 2),
                "rainProbability": round(float(parse_number(row.get("Rain %")) or 0), 2),
                "wind": round(float(parse_number(row.get("Wind km/h")) or 0), 2),
            }
            if risk_row["date"]:
                forecast_risk_data.append(risk_row)

    historical_summary_data = []
    for date, grp in df.groupby("Date", sort=True):
        summary_row = {"date": date}
        if temp_col:
            summary_row["minTemp"] = round_or_none(grp[temp_col].min())
            summary_row["maxTemp"] = round_or_none(grp[temp_col].max())
        else:
            summary_row["minTemp"] = None
            summary_row["maxTemp"] = None
        summary_row["precipitation"] = round_or_none(grp[precip_col].sum()) if precip_col else None
        summary_row["solarRadiation"] = round_or_none(grp[solar_col].mean()) if solar_col else None
        summary_row["windSpeed"] = round_or_none(grp[wind_col].mean()) if wind_col else None
        historical_summary_data.append(summary_row)

    forecast_table_data = forecast_table_rows(forecast)
    second_week_forecast_table_data = forecast_table_rows(second_week_forecast)

    js_text = app_js_path.read_text(encoding="utf-8")
    js_text = replace_js_const_array(js_text, "precipData", precip_data)
    js_text = replace_js_const_array(js_text, "tempData", temp_data)
    js_text = replace_js_const_array(js_text, "forecastRiskData", forecast_risk_data)
    js_text = replace_js_const_array(js_text, "historicalSummaryData", historical_summary_data)
    js_text = replace_js_const_array(js_text, "forecastTableData", forecast_table_data)
    js_text = replace_js_const_array(js_text, "secondWeekForecastTableData", second_week_forecast_table_data)
    js_text = replace_js_const_string(js_text, "dashboardUpdatedAt", datetime.now().strftime("%Y-%m-%d %H:%M"))
    app_js_path.write_text(js_text, encoding="utf-8")
    print(f"  ✔  Dashboard data updated → {app_js_path}")
    if historical_summary_data:
        print(
            f"     Dashboard period: {historical_summary_data[0]['date']} → "
            f"{historical_summary_data[-1]['date']} ({len(historical_summary_data)} days)"
        )


# ══════════════════════════════════════════════════════════════════════════════
#  EXCEL EXPORT
# ══════════════════════════════════════════════════════════════════════════════

def export_excel(df_out: pd.DataFrame, out_path: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Weather Data"

    header_font  = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    header_fill  = PatternFill("solid", start_color="1F4E79")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    data_font    = Font(name="Arial", size=10)
    thin         = Side(style="thin", color="BFBFBF")
    border       = Border(left=thin, right=thin, top=thin, bottom=thin)
    light_blue   = PatternFill("solid", start_color="DCE6F1")

    for col_idx, col_name in enumerate(df_out.columns, start=1):
        cell           = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = header_align
        cell.border    = border
    ws.row_dimensions[1].height = 30

    for row_idx, row in enumerate(df_out.itertuples(index=False), start=2):
        for col_idx, value in enumerate(row, start=1):
            cell        = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font   = data_font
            cell.border = border
            cell.alignment = Alignment(horizontal="left" if col_idx == 1 else "center")
            if col_idx > 1:
                cell.number_format = "0.0000"

    for row_idx in range(2, ws.max_row + 1, 2):
        for col_idx in range(1, len(df_out.columns) + 1):
            ws.cell(row=row_idx, column=col_idx).fill = light_blue

    ws.column_dimensions["A"].width = 18
    for col_idx in range(2, len(df_out.columns) + 1):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = max(14, len(df_out.columns[col_idx - 1]) + 2)

    ws.freeze_panes    = "A2"
    ws.auto_filter.ref = ws.dimensions
    wb.save(out_path)


# ══════════════════════════════════════════════════════════════════════════════
#  EMAIL  —  HTML builders
# ══════════════════════════════════════════════════════════════════════════════

# ── Shared style constants ─────────────────────────────────────────────────────
_HDR_STYLE  = "background:#1F4E79;color:white;padding:12px 16px;border-radius:4px;"
_TH_STYLE   = "background:#1F4E79;color:white;padding:8px 12px;text-align:left;"
_THC_STYLE  = "background:#1F4E79;color:white;padding:8px 12px;text-align:center;"
_TD_STYLE   = "padding:6px 12px;border:1px solid #ddd;"
_TDC_STYLE  = "padding:6px 12px;border:1px solid #ddd;text-align:center;"
_TBL_STYLE  = "width:100%;border-collapse:collapse;margin-bottom:20px;"
_EVEN_ROW   = "background:#EBF3FB;"


def _daily_summary_html(df_summary: pd.DataFrame) -> str:
    """Build an HTML table for the Zentra daily summary block."""
    if df_summary.empty:
        return "<p><i>No daily summary data available.</i></p>"

    header_html = ""
    for i, col in enumerate(df_summary.columns):
        style = _TH_STYLE if i == 0 else _THC_STYLE
        header_html += f'<th style="{style}">{col}</th>'

    rows_html = ""
    for i, (_, row) in enumerate(df_summary.iterrows()):
        bg = f'style="{_EVEN_ROW}"' if i % 2 == 0 else ""
        cells_html = ""
        for j, col in enumerate(df_summary.columns):
            style = _TD_STYLE if j == 0 else _TDC_STYLE
            value = row[col]
            if pd.isna(value):
                value = ""
            cells_html += f'<td style="{style}">{value}</td>'
        rows_html += f"""
        <tr {bg}>
          {cells_html}
        </tr>"""

    return f"""
    <h3 style="color:#1F4E79;">📊 7-Day Historical Measurements (ZentraCloud)</h3>
    <table style="{_TBL_STYLE}">
      <tr>
        {header_html}
      </tr>
      {rows_html}
    </table>"""


def _forecast_html(
    forecast_rows: list[dict] | None,
    title: str = "7-Day Weather Forecast (Open-Meteo)",
) -> str:
    """Build an HTML table for the Open-Meteo 7-day forecast block."""
    if not forecast_rows:
        return f"""
        <h3 style="color:#1F4E79;">🌤 {title}</h3>
        <p style="color:#C00000;"><i>Forecast data could not be retrieved.</i></p>"""

    rows_html = ""
    for i, row in enumerate(forecast_rows):
        bg = f'style="{_EVEN_ROW}"' if i % 2 == 0 else ""
        rows_html += f"""
        <tr {bg}>
          <td style="{_TD_STYLE}">{row["Date"]}</td>
          <td style="{_TD_STYLE}">{row["Condition"]}</td>
          <td style="{_TDC_STYLE}">{row["Max °C"]}</td>
          <td style="{_TDC_STYLE}">{row["Min °C"]}</td>
          <td style="{_TDC_STYLE}">{row["Rain mm"]}</td>
          <td style="{_TDC_STYLE}">{row["Rain %"]}</td>
          <td style="{_TDC_STYLE}">{row["Wind km/h"]}</td>
        </tr>"""

    return f"""
    <h3 style="color:#1F4E79;">🌤 {title}</h3>
    <p style="font-size:12px;color:#555;">
      Location: {LATITUDE}°N, {LONGITUDE}°E &nbsp;|&nbsp;
      Source: <a href="https://open-meteo.com/">open-meteo.com</a>
    </p>
    <table style="{_TBL_STYLE}">
      <tr>
        <th style="{_TH_STYLE}">Date</th>
        <th style="{_TH_STYLE}">Condition</th>
        <th style="{_THC_STYLE}">Max (°C)</th>
        <th style="{_THC_STYLE}">Min (°C)</th>
        <th style="{_THC_STYLE}">Rain (mm)</th>
        <th style="{_THC_STYLE}">Rain %</th>
        <th style="{_THC_STYLE}">Wind (km/h)</th>
      </tr>
      {rows_html}
    </table>"""


def _charts_html(chart_paths: dict[str, str] | None) -> str:
    """Build the inline chart section for the email."""
    if not chart_paths:
        return ""

    charts = []
    if "precip_chart" in chart_paths:
        charts.append("""
        <h3 style="color:#1F4E79;">Precipitation: 7-Day Historical and 14-Day Forecast</h3>
        <img src="cid:precip_chart" alt="Precipitation chart"
             style="width:100%;max-width:680px;border:1px solid #ddd;">""")
    if "temp_chart" in chart_paths:
        charts.append("""
        <h3 style="color:#1F4E79;">Min and Max Temperature: 7-Day Historical and 14-Day Forecast</h3>
        <img src="cid:temp_chart" alt="Temperature chart"
             style="width:100%;max-width:680px;border:1px solid #ddd;">""")

    if not charts:
        return ""

    return f"""
    <!-- ── Charts ── -->
    <div style="margin-top:8px;margin-bottom:20px;">
      {''.join(charts)}
    </div>"""


def build_email_body(
    start_date: str,
    end_date: str,
    quality_report: dict,
    total_rows: int,
    excel_filename: str,
    df_summary: pd.DataFrame = None,
    forecast_rows: list[dict] | None = None,
    second_week_forecast_rows: list[dict] | None = None,
    error: str = None,
    delayed_start: bool = False,
    delayed_hours: int = 0,
    delayed_start_str: str = None,
    chart_paths: dict[str, str] | None = None,
) -> str:
    run_time = datetime.now().strftime("%Y-%m-%d %H:%M")

    # ── Error fallback ────────────────────────────────────────────────────────
    if error:
        return f"""
        <html><body style="font-family:Arial,sans-serif;color:#333;">
        <h2 style="color:#C00000;">⚠ Zentra Cloud Weekly Run Failed</h2>
        <p><b>Run time:</b> {run_time}</p>
        <p><b>Device:</b> {DEVICE_SN} — Port {PORT_NUM}</p>
        <p><b>Period:</b> {start_date} → {end_date}</p>
        <hr>
        <p style="color:#C00000;"><b>Error:</b><br><pre>{error}</pre></p>
        </body></html>
        """

    # ── Daily summary & forecast blocks ──────────────────────────────────────
    summary_block  = _daily_summary_html(df_summary if df_summary is not None else pd.DataFrame())
    combined_forecast_rows = [
        *(forecast_rows or []),
        *(second_week_forecast_rows or []),
    ]
    forecast_block = _forecast_html(combined_forecast_rows, "14-Day Forecast (Open-Meteo)")
    charts_block   = _charts_html(chart_paths)

    return f"""
    <html>
    <body style="font-family:Arial,sans-serif;color:#333;max-width:700px;margin:auto;">

      <!-- ── Header ── -->
      <h2 style="{_HDR_STYLE}">
        📡 Zentra Cloud — Weekly Weather Data Report
      </h2>

      <!-- ── Run metadata ── -->
      <table style="width:100%;margin-bottom:16px;">
        <tr><td><b>Run time</b></td><td>{run_time}</td></tr>
        <tr><td><b>Device</b></td><td>{DEVICE_SN} — Port {PORT_NUM}</td></tr>
        <tr><td><b>Period</b></td><td>{start_date} → {end_date}</td></tr>
        <tr><td><b>Hourly records</b></td><td>{total_rows:,}</td></tr>
        <tr><td><b>Excel file</b></td><td>{excel_filename}</td></tr>
        <tr><td><b>Location (forecast)</b></td><td>{LATITUDE}°N, {LONGITUDE}°E</td></tr>
      </table>

      <!-- ── 7-day historical daily summary (Zentra) ── -->
      {summary_block}

      <!-- ── 14-day forecast (Open-Meteo) ── -->
      {forecast_block}

      {charts_block}

      <p style="margin-top:24px;font-size:12px;color:#888;">
        The full hourly Excel file is saved to
        your OneDrive ZentraCloud folder.<br>
        Forecast data provided by <a href="https://open-meteo.com/">Open-Meteo</a>.
      </p>

    </body>
    </html>
    """


# ══════════════════════════════════════════════════════════════════════════════
#  EMAIL  —  Send
# ══════════════════════════════════════════════════════════════════════════════

def send_email(
    subject: str,
    html_body: str,
    inline_images: dict[str, str] | None = None,
) -> None:
    recipients = [addr.strip() for addr in EMAIL_RECIPIENT if str(addr).strip()]
    if not recipients:
        print("  ⚠  No email recipients configured; skipping email send.")
        return

    msg             = MIMEMultipart("related")
    msg["Subject"]  = subject
    msg["From"]     = GMAIL_SENDER
    msg["To"]       = ", ".join(recipients)

    alternative = MIMEMultipart("alternative")
    alternative.attach(MIMEText(html_body, "html"))
    msg.attach(alternative)

    for cid, image_path in (inline_images or {}).items():
        if not image_path or not os.path.exists(image_path):
            continue
        with open(image_path, "rb") as f:
            image = MIMEImage(f.read())
        image.add_header("Content-ID", f"<{cid}>")
        image.add_header("Content-Disposition", "inline", filename=os.path.basename(image_path))
        msg.attach(image)

    with smtplib.SMTP_SSL("smtp.gmail.com", 465) as server:
        server.login(GMAIL_SENDER, GMAIL_APP_PASSWORD)
        server.sendmail(GMAIL_SENDER, recipients, msg.as_string())

    print(f"  ✔  Email sent to: {', '.join(recipients)}")


# ══════════════════════════════════════════════════════════════════════════════
#  MAIN
# ══════════════════════════════════════════════════════════════════════════════

def main(dashboard_only: bool = False):
    start_date, end_date = historical_range()
    fetch_end_date = zentra_fetch_end_date()

    print("=" * 62)
    print("  Zentra Cloud – Weekly Automated Runner (+ Open-Meteo)")
    print("=" * 62)
    print(f"  Device  : {DEVICE_SN}  |  Port: {PORT_NUM}")
    print(f"  Run date: {datetime.now().strftime('%Y-%m-%d %H:%M')}\n")

    try:
        # ── 1. Fetch Zentra data ──────────────────────────────────────────────
        print("── Step 1: Fetching Zentra Cloud data ───────────────────────────")
        df_raw = fetch_all(start_date, fetch_end_date)
        if df_raw.empty:
            raise ValueError("No data returned from API for the given period.")

        # ── 2. Process hourly ─────────────────────────────────────────────────
        print("\n── Step 2: Processing data ──────────────────────────────────────")
        df_out, quality_report, delayed_start, delayed_hours, delayed_start_str = process(df_raw)
        df_out = filter_output_to_date_range(df_out, start_date, end_date)
        if df_out.empty:
            raise ValueError("No processed data remained after filtering to the historical date range.")
        print(f"  Period  : {start_date}  →  {end_date}  (requested)")
        if fetch_end_date != end_date:
            print(f"  API end : {fetch_end_date}  (fetch boundary)")
        print(f"  Data    : {df_out['Datetime'].iloc[0]}  →  {df_out['Datetime'].iloc[-1]}  (actual)")
        data_dates = sorted(pd.to_datetime(df_out["Datetime"], errors="coerce").dt.strftime("%Y-%m-%d").dropna().unique())
        print(f"  Dates   : {', '.join(data_dates)}")
        if end_date not in data_dates:
            print(f"  ⚠  Expected end date {end_date} was not returned by Zentra after processing.")

        # ── 3. Build daily summary ────────────────────────────────────────────
        print("\n── Step 3: Building daily summary ───────────────────────────────")
        df_summary = build_daily_summary(df_out)
        print(f"  ✔  Daily summary: {len(df_summary)} daily rows with "
              f"{max(len(df_summary.columns) - 1, 0)} variables.")

        # ── 4. Fetch Open-Meteo forecast ──────────────────────────────────────
        print("\n── Step 4: Fetching Open-Meteo forecast ─────────────────────────")
        from datetime import datetime as dt, timedelta as td
        forecast_start = dt.now().date().strftime("%Y-%m-%d")   # forecast from today
        second_week_forecast_start = (dt.now().date() + td(days=7)).strftime("%Y-%m-%d")
        forecast_rows = fetch_forecast(forecast_start)
        second_week_forecast_rows = fetch_forecast(second_week_forecast_start)

        # ── 5. Update static dashboard data ──────────────────────────────────
        print("\n── Step 5: Updating dashboard data ──────────────────────────────")
        update_dashboard_app_js(df_out, forecast_rows, second_week_forecast_rows)
        ndvi_meta = fetch_latest_sentinel2_ndvi()
        update_dashboard_ndvi_metadata(ndvi_meta)

        if dashboard_only:
            print("\n── Dashboard-only mode complete; Excel export and email skipped. ──")
            return

        # ── 6. Export to Excel ────────────────────────────────────────────────
        print("\n── Step 6: Exporting Excel ──────────────────────────────────────")
        os.makedirs(OUT_DIR, exist_ok=True)
        ts       = datetime.now().strftime("%Y%m%d_%H%M%S")
        port_tag = f"_port{PORT_NUM}"
        filename = f"zentra_weather_{DEVICE_SN}{port_tag}_{start_date}_{end_date}_hourly_{ts}.xlsx"
        out_path = os.path.join(OUT_DIR, filename)
        export_excel(df_out, out_path)
        print(f"  ✔  Excel saved → {out_path}")

        # ── 6. Build email charts ────────────────────────────────────────────
        print("\n── Step 7: Building email charts ────────────────────────────────")
        chart_paths = build_email_charts(df_out, forecast_rows, second_week_forecast_rows, OUT_DIR, ts)
        print(f"  ✔  Email charts: {len(chart_paths)} generated.")

        # ── 7. Send email ─────────────────────────────────────────────────────
        print("\n── Step 8: Sending email report ─────────────────────────────────")
        subject   = f"Zentra Weather Data + Forecast — {start_date} to {end_date}"
        html_body = build_email_body(
            start_date        = start_date,
            end_date          = end_date,
            quality_report    = quality_report,
            total_rows        = len(df_out),
            excel_filename    = filename,
            df_summary        = df_summary,
            forecast_rows     = forecast_rows,
            second_week_forecast_rows = second_week_forecast_rows,
            delayed_start     = delayed_start,
            delayed_hours     = delayed_hours,
            delayed_start_str = delayed_start_str,
            chart_paths       = chart_paths,
        )
        send_email(subject, html_body, inline_images=chart_paths)

    except Exception as e:
        err_text = traceback.format_exc()
        print(f"\n✗ Run failed:\n{err_text}")
        try:
            subject   = f"⚠ Zentra Weekly Run FAILED — {start_date} to {end_date}"
            html_body = build_email_body(start_date, end_date, {}, 0, "", error=err_text)
            send_email(subject, html_body)
        except Exception as email_err:
            print(f"  ✗ Could not send failure email: {email_err}")


if __name__ == "__main__":
    main(dashboard_only="--dashboard-only" in sys.argv)
